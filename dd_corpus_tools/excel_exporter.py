#!/usr/bin/env python3
"""
Consolidated Excel Export
Creates a single comprehensive CSV and Excel workbook optimized for pivot table analysis
"""

import os
from pathlib import Path
from datetime import datetime
from collections import defaultdict
import csv

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

from .analyzer import NewspaperCorpusAnalyzer
from .constants import NEWSPAPER_NAMES, NEWSPAPER_METADATA


class ConsolidatedExporter(NewspaperCorpusAnalyzer):
    """Exports consolidated data optimized for Excel pivot tables"""
    
    def __init__(self, corpus_root):
        super().__init__(corpus_root)
        self.issues_detail = defaultdict(lambda: defaultdict(list))
        self.frequency_cache = {}
    
    def scan_corpus(self):
        """Extended scan that tracks page details"""
        super().scan_corpus()
        
        # Build detailed issue information
        for newspaper in self.data:
            for year in self.data[newspaper]:
                for page_info in self.data[newspaper][year]:
                    date = page_info['date']
                    self.issues_detail[newspaper][date].append(page_info['page'])
        
        # Calculate frequencies
        for newspaper in self.issues:
            self._calculate_frequency(newspaper)
    
    def _calculate_frequency(self, newspaper):
        """Calculate publication frequency for a newspaper"""
        dates = sorted(self.issues[newspaper])
        
        if len(dates) < 2:
            self.frequency_cache[newspaper] = {
                'frequency': 'Unknown',
                'avg_gap': 0,
                'expected_gap': 30
            }
            return
        
        date_objects = [datetime.strptime(d, '%Y-%m-%d') for d in dates]
        gaps = [(date_objects[i+1] - date_objects[i]).days 
               for i in range(len(date_objects)-1)]
        avg_gap = sum(gaps) / len(gaps)
        
        if avg_gap <= 1.5:
            freq, expected = "Daily", 1
        elif avg_gap <= 4:
            freq, expected = "2-3x per week", 3
        elif avg_gap <= 8:
            freq, expected = "Weekly", 7
        elif avg_gap <= 16:
            freq, expected = "Bi-weekly", 14
        else:
            freq, expected = "Monthly/Irregular", 30
        
        self.frequency_cache[newspaper] = {
            'frequency': freq,
            'avg_gap': avg_gap,
            'expected_gap': expected
        }
    
    def export_consolidated_csv(self, output_file='corpus_pivot_table_data.csv'):
        """Export consolidated issue-level data for pivot table analysis"""
        # Create output directory
        output_dir = Path('dd_corpus_reports')
        output_dir.mkdir(exist_ok=True)
        output_path = output_dir / output_file
        
        print(f"Generating consolidated CSV for pivot table analysis...")
        
        with open(output_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            
            # Comprehensive header for pivot table analysis
            writer.writerow([
                'Date',
                'Year',
                'Month',
                'Day',
                'Quarter',
                'Decade',
                'Newspaper_Code',
                'Newspaper_Name',
                'Region',
                'Publisher_Type',
                'Administrative_Level',
                'Province',
                'Publication_Type',
                'Pages_In_Issue',
                'Has_Missing_Pages',
                'Has_Duplicate_Pages',
                'Primary_Library',
                'All_Libraries',
                'Num_Libraries',
                'Estimated_Frequency',
                'Avg_Gap_Days',
                'Is_Complete_Issue',
                'Completeness_Pct'
            ])
            
            # Write data for each issue
            for newspaper in sorted(self.issues_detail.keys()):
                metadata = NEWSPAPER_METADATA.get(newspaper, {})
                freq_info = self.frequency_cache.get(newspaper, {})
                
                for date in sorted(self.issues_detail[newspaper].keys()):
                    pages = sorted(self.issues_detail[newspaper][date])
                    
                    # Parse date
                    year, month, day = map(int, date.split('-'))
                    quarter = f"Q{(month - 1) // 3 + 1}"
                    decade = f"{(year // 10) * 10}s"
                    
                    # Check completeness
                    max_page = max(pages)
                    expected_pages = set(range(1, max_page + 1))
                    missing_pages = expected_pages - set(pages)
                    has_missing = len(missing_pages) > 0
                    has_duplicates = len(pages) != len(set(pages))
                    completeness_pct = (len(set(pages)) / max_page * 100) if max_page > 0 else 100
                    is_complete = not has_missing and not has_duplicates
                    
                    # Get libraries for this issue
                    issue_libraries = set()
                    for page_data in self.data[newspaper][year]:
                        if page_data['date'] == date:
                            issue_libraries.add(page_data['library'])
                    
                    primary_library = sorted(issue_libraries)[0] if issue_libraries else ''
                    all_libraries = '; '.join(sorted(issue_libraries))
                    
                    writer.writerow([
                        date,
                        year,
                        month,
                        day,
                        quarter,
                        decade,
                        newspaper,
                        NEWSPAPER_NAMES.get(newspaper, newspaper),
                        metadata.get('region', ''),
                        metadata.get('publisher', ''),
                        metadata.get('level', ''),
                        metadata.get('province', ''),
                        metadata.get('type', ''),
                        len(pages),
                        'Yes' if has_missing else 'No',
                        'Yes' if has_duplicates else 'No',
                        primary_library,
                        all_libraries,
                        len(issue_libraries),
                        freq_info.get('frequency', ''),
                        round(freq_info.get('avg_gap', 0), 1),
                        'Yes' if is_complete else 'No',
                        round(completeness_pct, 1)
                    ])
        
        print(f"✓ Consolidated CSV created: {output_path.absolute()}")
        print(f"  Total issues: {sum(len(dates) for dates in self.issues_detail.values())}")
        print(f"  Ready for Excel pivot table analysis!")
        print()
        return output_path
    
    def export_excel_workbook(self, output_file='corpus_analysis_workbook.xlsx'):
        """Export comprehensive Excel workbook with multiple sheets"""
        
        if not EXCEL_AVAILABLE:
            print("Warning: openpyxl not installed. Install with: pip install openpyxl")
            print("Skipping Excel workbook generation.")
            return None
        
        # Create output directory
        output_dir = Path('dd_corpus_reports')
        output_dir.mkdir(exist_ok=True)
        output_path = output_dir / output_file
        
        print(f"Generating Excel workbook with multiple sheets...")
        
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Sheet 1: Issue-level data (main pivot table data)
        self._create_issues_sheet(wb)
        
        # Sheet 2: Newspaper summary
        self._create_newspapers_sheet(wb)
        
        # Sheet 3: Library holdings
        self._create_libraries_sheet(wb)
        
        # Sheet 4: Yearly statistics
        self._create_yearly_sheet(wb)
        
        # Sheet 5: Missing issues summary
        self._create_missing_issues_sheet(wb)
        
        # Sheet 6: Completeness statistics
        self._create_completeness_sheet(wb)
        
        wb.save(output_path)
        
        print(f"✓ Excel workbook created: {output_path.absolute()}")
        print(f"  6 sheets with comprehensive analysis")
        print(f"  Ready for pivot tables and charts!")
        print()
        return output_path
    
    def _create_issues_sheet(self, wb):
        """Create the main issues sheet"""
        ws = wb.create_sheet("Issues_Data")
        
        # Header
        headers = [
            'Date', 'Year', 'Month', 'Day', 'Quarter', 'Decade',
            'Newspaper_Code', 'Newspaper_Name', 'Region', 'Publisher_Type',
            'Administrative_Level', 'Province', 'Publication_Type',
            'Pages_In_Issue', 'Has_Missing_Pages', 'Has_Duplicate_Pages',
            'Primary_Library', 'All_Libraries', 'Num_Libraries',
            'Estimated_Frequency', 'Avg_Gap_Days', 'Is_Complete_Issue',
            'Completeness_Pct'
        ]
        
        ws.append(headers)
        
        # Format header
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        # Data rows
        for newspaper in sorted(self.issues_detail.keys()):
            metadata = NEWSPAPER_METADATA.get(newspaper, {})
            freq_info = self.frequency_cache.get(newspaper, {})
            
            for date in sorted(self.issues_detail[newspaper].keys()):
                pages = sorted(self.issues_detail[newspaper][date])
                year, month, day = map(int, date.split('-'))
                quarter = f"Q{(month - 1) // 3 + 1}"
                decade = f"{(year // 10) * 10}s"
                
                max_page = max(pages)
                expected_pages = set(range(1, max_page + 1))
                missing_pages = expected_pages - set(pages)
                has_missing = len(missing_pages) > 0
                has_duplicates = len(pages) != len(set(pages))
                completeness_pct = (len(set(pages)) / max_page * 100) if max_page > 0 else 100
                is_complete = not has_missing and not has_duplicates
                
                issue_libraries = set()
                for page_data in self.data[newspaper][year]:
                    if page_data['date'] == date:
                        issue_libraries.add(page_data['library'])
                
                primary_library = sorted(issue_libraries)[0] if issue_libraries else ''
                all_libraries = '; '.join(sorted(issue_libraries))
                
                ws.append([
                    date, year, month, day, quarter, decade,
                    newspaper, NEWSPAPER_NAMES.get(newspaper, newspaper),
                    metadata.get('region', ''), metadata.get('publisher', ''),
                    metadata.get('level', ''), metadata.get('province', ''),
                    metadata.get('type', ''), len(pages),
                    'Yes' if has_missing else 'No',
                    'Yes' if has_duplicates else 'No',
                    primary_library, all_libraries, len(issue_libraries),
                    freq_info.get('frequency', ''),
                    round(freq_info.get('avg_gap', 0), 1),
                    'Yes' if is_complete else 'No',
                    round(completeness_pct, 1)
                ])
        
        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_newspapers_sheet(self, wb):
        """Create newspaper summary sheet"""
        ws = wb.create_sheet("Newspapers_Summary")
        
        headers = [
            'Newspaper_Code', 'Newspaper_Name', 'Region', 'Publisher_Type',
            'Administrative_Level', 'Province', 'Total_Issues', 'Total_Pages',
            'First_Year', 'Last_Year', 'Years_Covered', 'Estimated_Frequency',
            'Libraries_Count', 'Libraries_List'
        ]
        ws.append(headers)
        
        # Format header
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        for newspaper in sorted(self.data.keys()):
            metadata = NEWSPAPER_METADATA.get(newspaper, {})
            freq_info = self.frequency_cache.get(newspaper, {})
            years = sorted(self.data[newspaper].keys())
            
            libraries_with_this = [lib for lib in self.libraries 
                                  if newspaper in self.libraries[lib]]
            
            ws.append([
                newspaper,
                NEWSPAPER_NAMES.get(newspaper, newspaper),
                metadata.get('region', ''),
                metadata.get('publisher', ''),
                metadata.get('level', ''),
                metadata.get('province', ''),
                len(self.issues[newspaper]),
                self.pages_by_newspaper[newspaper],
                min(years),
                max(years),
                len(years),
                freq_info.get('frequency', ''),
                len(libraries_with_this),
                ', '.join(sorted(libraries_with_this))
            ])
        
        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_libraries_sheet(self, wb):
        """Create library holdings sheet"""
        ws = wb.create_sheet("Library_Holdings")
        
        headers = [
            'Library_Code', 'Newspaper_Code', 'Newspaper_Name',
            'Total_Issues', 'Total_Pages', 'First_Year', 'Last_Year',
            'Years_List'
        ]
        ws.append(headers)
        
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        for library in sorted(self.libraries.keys()):
            for newspaper in sorted(self.libraries[library].keys()):
                years = sorted(self.library_years[library][newspaper])
                
                ws.append([
                    library,
                    newspaper,
                    NEWSPAPER_NAMES.get(newspaper, newspaper),
                    len(self.library_issues[library][newspaper]),
                    self.libraries[library][newspaper],
                    min(years) if years else '',
                    max(years) if years else '',
                    ', '.join(map(str, years))
                ])
        
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_yearly_sheet(self, wb):
        """Create yearly statistics sheet"""
        ws = wb.create_sheet("Yearly_Statistics")
        
        headers = ['Year', 'Total_Issues', 'Total_Pages', 'Newspapers_Count', 'Newspapers_List']
        ws.append(headers)
        
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        for year in sorted(self.pages_by_year.keys()):
            total_issues = sum(len(self.issues_by_year[np][year]) 
                             for np in self.issues_by_year if year in self.issues_by_year[np])
            newspapers_active = [np for np in self.data if year in self.data[np]]
            
            ws.append([
                year,
                total_issues,
                self.pages_by_year[year],
                len(newspapers_active),
                ', '.join(sorted(newspapers_active))
            ])
        
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_missing_issues_sheet(self, wb):
        """Create missing issues summary sheet"""
        ws = wb.create_sheet("Missing_Issues")
        
        headers = [
            'Newspaper_Code', 'Newspaper_Name', 'Issues_In_Corpus',
            'Estimated_Frequency', 'Avg_Gap_Days', 'Significant_Gaps',
            'Estimated_Missing', 'Completeness_Pct'
        ]
        ws.append(headers)
        
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        for newspaper in sorted(self.issues.keys()):
            dates = sorted(self.issues[newspaper])
            if len(dates) < 2:
                continue
            
            freq_info = self.frequency_cache[newspaper]
            date_objects = [datetime.strptime(d, '%Y-%m-%d') for d in dates]
            
            significant_gaps = 0
            total_estimated_missing = 0
            
            for i in range(len(date_objects) - 1):
                gap_days = (date_objects[i + 1] - date_objects[i]).days
                if gap_days > freq_info['expected_gap'] * 2:
                    significant_gaps += 1
                    missing_estimate = max(0, int(gap_days / freq_info['expected_gap']) - 1)
                    total_estimated_missing += missing_estimate
            
            completeness = len(dates) / (len(dates) + total_estimated_missing) * 100 if total_estimated_missing > 0 else 100
            
            ws.append([
                newspaper,
                NEWSPAPER_NAMES.get(newspaper, newspaper),
                len(dates),
                freq_info['frequency'],
                round(freq_info['avg_gap'], 1),
                significant_gaps,
                total_estimated_missing,
                round(completeness, 1)
            ])
        
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_completeness_sheet(self, wb):
        """Create issue completeness statistics sheet"""
        ws = wb.create_sheet("Issue_Completeness")
        
        headers = [
            'Newspaper_Code', 'Newspaper_Name', 'Total_Issues',
            'Complete_Issues', 'Issues_With_Missing_Pages',
            'Issues_With_Duplicates', 'Complete_Issues_Pct'
        ]
        ws.append(headers)
        
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        for newspaper in sorted(self.issues_detail.keys()):
            complete_count = 0
            missing_pages_count = 0
            duplicates_count = 0
            
            for date in self.issues_detail[newspaper]:
                pages = sorted(self.issues_detail[newspaper][date])
                max_page = max(pages)
                expected_pages = set(range(1, max_page + 1))
                missing_pages = expected_pages - set(pages)
                has_duplicates = len(pages) != len(set(pages))
                
                if not missing_pages and not has_duplicates:
                    complete_count += 1
                if missing_pages:
                    missing_pages_count += 1
                if has_duplicates:
                    duplicates_count += 1
            
            total_issues = len(self.issues_detail[newspaper])
            complete_pct = (complete_count / total_issues * 100) if total_issues > 0 else 0
            
            ws.append([
                newspaper,
                NEWSPAPER_NAMES.get(newspaper, newspaper),
                total_issues,
                complete_count,
                missing_pages_count,
                duplicates_count,
                round(complete_pct, 1)
            ])
        
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def export_all(self):
        """Export both CSV and Excel workbook"""
        print("\n" + "=" * 80)
        print("CONSOLIDATED EXPORT FOR EXCEL ANALYSIS")
        print("=" * 80 + "\n")
        
        self.scan_corpus()
        
        # Export CSV (always available)
        csv_file = self.export_consolidated_csv()
        
        # Export Excel workbook (if openpyxl available)
        excel_file = self.export_excel_workbook()
        
        print("=" * 80)
        print("EXPORT COMPLETE")
        print("=" * 80)
        print(f"\n✓ CSV file: {csv_file}")
        if excel_file:
            print(f"✓ Excel workbook: {excel_file}")
        else:
            print("  Excel workbook not created (install openpyxl for Excel export)")
        
        print("\nPivot Table Suggestions:")
        print("  • Pages by Year and Newspaper")
        print("  • Issues by Region and Publisher Type")
        print("  • Library Holdings by Administrative Level")
        print("  • Completeness by Decade and Province")
        print("  • Missing Issues by Publication Frequency")
        print()


if __name__ == '__main__':
    # For testing only
    import sys
    if len(sys.argv) >= 2:
        exporter = ConsolidatedExporter(sys.argv[1])
        exporter.export_all()
